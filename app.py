from flask import Flask, request, jsonify, send_file, render_template, Response
from pypdf import PdfReader, PdfWriter
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, re, uuid, zipfile, os, tempfile, gc
from datetime import datetime
from collections import Counter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024

# Sessions store temp file PATHS, not raw bytes
SESSIONS = {}

MONTHS = {
    'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
    'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12
}

# ─── Helpers ──────────────────────────────────────────────────────────────────

def parse_date(date_str):
    date_str = date_str.strip()
    m = re.match(r'(\d{1,2})[\s\-/]([A-Za-z]{3})[\s\-/](\d{2,4})', date_str)
    if m:
        d, mo, y = int(m.group(1)), MONTHS.get(m.group(2).lower(), 0), int(m.group(3))
        if y < 100: y += 2000
        if mo: return datetime(y, mo, d)
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', date_str)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return datetime(y, mo, d)
    return None

def parse_amount(amt_str):
    if not amt_str:
        return None
    s = str(amt_str).strip().replace(',', '').replace(' ', '')
    negative = s.startswith('(') or s.startswith('-')
    s = s.strip('()-').strip()
    try:
        val = float(s)
        return -val if negative else val
    except:
        return None

DATE_RE = re.compile(r'\b(\d{1,2}[\s\-/][A-Za-z]{3}[\s\-/]\d{2,4}|\d{1,2}/\d{1,2}/\d{4})\b')
AMOUNT_RE = re.compile(r'[\d,]+\.\d{2}')

# ─── Temp file helpers ────────────────────────────────────────────────────────

def save_to_tmp(data: bytes) -> str:
    fd, path = tempfile.mkstemp(suffix='.pdf')
    with os.fdopen(fd, 'wb') as f:
        f.write(data)
    return path

def delete_tmp(path: str):
    try:
        os.unlink(path)
    except Exception:
        pass

def cleanup_session(session_id):
    session = SESSIONS.pop(session_id, {})
    for bucket in ('unlocked', 'locked'):
        for path in session.get(bucket, {}).values():
            delete_tmp(path)
    if 'xlsx_path' in session:
        delete_tmp(session['xlsx_path'])

# ─── PDF parsing: file-path-based to avoid holding full bytes in RAM ──────────

def extract_text_streamed(pdf_path):
    """Extract text page-by-page. OCR fallback is also page-by-page at lower DPI."""
    pages_text = []
    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        for page in pdf.pages:
            t = page.extract_text() or ''
            pages_text.append(t)
            page.flush_cache()

    full_text = '\n'.join(pages_text)

    if len(full_text.strip()) < 100 * max(len(pages_text), 1):
        try:
            from pdf2image import convert_from_path
            import pytesseract
            ocr_pages = []
            for page_num in range(1, total_pages + 1):
                images = convert_from_path(pdf_path, dpi=150,
                                           first_page=page_num, last_page=page_num)
                if images:
                    ocr_pages.append(pytesseract.image_to_string(images[0]))
                    del images
                    gc.collect()
            full_text = '\n'.join(ocr_pages)
        except Exception:
            pass

    return full_text

def extract_tables_streamed(pdf_path):
    all_tables = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            all_tables.extend(tables)
            page.flush_cache()
    return all_tables

def detect_columns(header_row):
    mapping = {}
    for i, cell in enumerate(header_row):
        if not cell: continue
        h = str(cell).lower().strip()
        if any(w in h for w in ['date', 'txn date', 'trans date', 'transaction date', 'post date']):
            mapping.setdefault('date', i)
        elif any(w in h for w in ['desc', 'narration', 'particular', 'detail', 'merchant', 'transaction detail']):
            mapping.setdefault('description', i)
        elif 'debit' in h and 'credit' not in h:
            mapping.setdefault('debit', i)
        elif 'credit' in h and 'debit' not in h:
            mapping.setdefault('credit', i)
        elif any(w in h for w in ['amount', 'amt', 'inr']):
            mapping.setdefault('amount', i)
        elif 'type' in h or 'dr/cr' in h or 'cr/dr' in h:
            mapping.setdefault('type', i)
    return mapping

def parse_transactions_from_tables(tables):
    transactions = []
    for table in tables:
        if not table or len(table) < 2:
            continue
        header_idx = None
        for i, row in enumerate(table):
            if not row: continue
            row_text = ' '.join(str(c).lower() for c in row if c)
            if any(w in row_text for w in ['date', 'description', 'amount', 'debit', 'narration']):
                header_idx = i
                break
        if header_idx is None:
            continue
        col_map = detect_columns(table[header_idx])
        if 'date' not in col_map:
            continue
        for row in table[header_idx + 1:]:
            if not row or all(not c for c in row):
                continue
            date_str = str(row[col_map['date']] or '').strip()
            parsed_date = parse_date(date_str)
            if not parsed_date:
                continue
            desc = str(row[col_map.get('description', 1)] or '').strip() if col_map.get('description') is not None else ''
            amount = None
            txn_type = ''
            if 'debit' in col_map and 'credit' in col_map:
                debit_val = parse_amount(row[col_map['debit']])
                credit_val = parse_amount(row[col_map['credit']])
                if debit_val:
                    amount = debit_val; txn_type = 'Debit'
                elif credit_val:
                    amount = credit_val; txn_type = 'Credit'
            elif 'amount' in col_map:
                amount = parse_amount(row[col_map['amount']])
                if 'type' in col_map:
                    txn_type = str(row[col_map['type']] or '').strip()
                else:
                    txn_type = 'Credit' if (amount and amount < 0) else 'Debit'
                    if amount: amount = abs(amount)
            if amount is None:
                continue
            transactions.append({
                'date': parsed_date,
                'description': desc,
                'amount': abs(amount),
                'type': txn_type or ('Credit' if amount < 0 else 'Debit'),
            })
    return transactions

def parse_transactions_from_text(text):
    transactions = []
    for line in text.split('\n'):
        line = line.strip()
        dates = DATE_RE.findall(line)
        if not dates: continue
        amounts = AMOUNT_RE.findall(line)
        if not amounts: continue
        parsed_date = parse_date(dates[0])
        if not parsed_date: continue
        desc = line
        for d in dates: desc = desc.replace(d, '')
        for a in amounts: desc = desc.replace(a, '').replace(',', '')
        txn_type = 'Credit' if re.search(r'\b(cr|credit|payment|reversal)\b', desc, re.I) else 'Debit'
        desc = re.sub(r'\b(debit|credit|dr|cr)\b', '', desc, flags=re.I).strip(' -|/')
        desc = re.sub(r'\s{2,}', ' ', desc).strip()
        amount = parse_amount(amounts[-1])
        if not amount: continue
        transactions.append({'date': parsed_date, 'description': desc,
                             'amount': abs(amount), 'type': txn_type})
    return transactions

def extract_account_info(text):
    info = {}
    m = re.search(r'(?:account|card)[\s\w]*?:?\s*([Xx*\d]{4}[\s\-]?[Xx*\d]{4}[\s\-]?[Xx*\d]{4}[\s\-]?[\dXx*]{4})', text, re.I)
    if m: info['card_number'] = m.group(1).strip()
    m = re.search(r'(?:period|from|statement date)[:\s]+([A-Za-z\d\s]+?)\s+to\s+([A-Za-z\d\s]+?)(?:\n|$)', text, re.I)
    if m:
        info['period_from'] = m.group(1).strip()
        info['period_to'] = m.group(2).strip()
    return info

def parse_pdf_from_path(pdf_path, source_name=''):
    """Parse a PDF given a file path — no full bytes held in RAM."""
    text = extract_text_streamed(pdf_path)
    tables = extract_tables_streamed(pdf_path)
    account_info = extract_account_info(text)
    account_info['source'] = source_name

    transactions = parse_transactions_from_tables(tables)
    if len(transactions) < 2:
        transactions = parse_transactions_from_text(text)

    del text, tables
    gc.collect()

    for t in transactions:
        t['source'] = source_name
    return account_info, transactions

# ─── Excel builder ────────────────────────────────────────────────────────────

def build_excel(all_transactions, account_infos):
    wb = openpyxl.Workbook()

    HDR_FILL = PatternFill('solid', start_color='1F3864')
    ALT_FILL = PatternFill('solid', start_color='EBF0FA')
    DEB_FILL = PatternFill('solid', start_color='FFF0F0')
    CRD_FILL = PatternFill('solid', start_color='F0FFF4')
    SUM_FILL = PatternFill('solid', start_color='FFF8E1')
    HDR_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    BODY_FONT = Font(name='Arial', size=10)
    BOLD_FONT = Font(name='Arial', bold=True, size=10)
    thin = Side(style='thin', color='D0D0D0')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(ws, row, col, val, width=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border
        if width: ws.column_dimensions[get_column_letter(col)].width = width

    ws = wb.active
    ws.title = 'All Transactions'
    ws.freeze_panes = 'A3'
    ws.merge_cells('A1:G1')
    title = ws['A1']
    title.value = 'SBI Card — Consolidated Statement'
    title.font = Font(name='Arial', bold=True, size=14, color='1F3864')
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    headers = ['#', 'Date', 'Description', 'Type', 'Debit (₹)', 'Credit (₹)', 'Source File']
    widths  = [5,  14,    45,             12,     16,           16,            30]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        hdr_cell(ws, 2, col, h, w)
    ws.row_dimensions[2].height = 22

    sorted_txns = sorted(all_transactions, key=lambda x: x['date'])

    for i, txn in enumerate(sorted_txns, 1):
        row = i + 2
        is_credit = txn['type'].lower() in ('credit', 'cr', 'payment', 'reversal')
        fill = CRD_FILL if is_credit else (ALT_FILL if i % 2 == 0 else PatternFill())
        vals = [i, txn['date'].strftime('%d %b %Y'), txn['description'], txn['type'],
                None if is_credit else txn['amount'],
                txn['amount'] if is_credit else None,
                txn['source']]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = BODY_FONT; c.fill = fill; c.border = thin_border
            if col == 1:
                c.alignment = Alignment(horizontal='center')
            elif col in (5, 6) and val is not None:
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right')

    total_row = len(sorted_txns) + 3
    for col in range(1, 8):
        ws.cell(row=total_row, column=col).fill = SUM_FILL
        ws.cell(row=total_row, column=col).border = thin_border
    ws.cell(row=total_row, column=3, value='TOTAL').font = BOLD_FONT

    for col, formula in [(5, f'=SUM(E3:E{total_row-1})'), (6, f'=SUM(F3:F{total_row-1})')]:
        c = ws.cell(row=total_row, column=col, value=formula)
        c.font = BOLD_FONT; c.fill = SUM_FILL
        c.number_format = '#,##0.00'
        c.alignment = Alignment(horizontal='right')
        c.border = thin_border

    # Summary sheet
    ws2 = wb.create_sheet('Summary')
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 22
    ws2.merge_cells('A1:B1')
    t2 = ws2['A1']
    t2.value = 'Statement Summary'
    t2.font = Font(name='Arial', bold=True, size=13, color='1F3864')
    t2.alignment = Alignment(horizontal='center')
    ws2.row_dimensions[1].height = 28

    def sum_row(row, label, val, fmt='#,##0.00'):
        c1 = ws2.cell(row=row, column=1, value=label)
        c2 = ws2.cell(row=row, column=2, value=val)
        c1.font = BODY_FONT; c1.border = thin_border
        c2.font = BODY_FONT; c2.border = thin_border
        c2.number_format = fmt
        c2.alignment = Alignment(horizontal='right')
        return c1, c2

    def bold_row(row, label, val, fill=None, fmt='#,##0.00'):
        c1, c2 = sum_row(row, label, val, fmt)
        c1.font = BOLD_FONT; c2.font = BOLD_FONT
        if fill: c1.fill = fill; c2.fill = fill

    r = 2
    hdr_cell(ws2, r, 1, 'Metric'); hdr_cell(ws2, r, 2, 'Value'); r += 1
    sum_row(r, 'Total Transactions', len(sorted_txns), fmt='0'); r += 1
    sum_row(r, 'Date Range From', sorted_txns[0]['date'].strftime('%d %b %Y') if sorted_txns else '-', fmt='@'); r += 1
    sum_row(r, 'Date Range To',   sorted_txns[-1]['date'].strftime('%d %b %Y') if sorted_txns else '-', fmt='@'); r += 1
    sum_row(r, 'Source Files', len(account_infos), fmt='0'); r += 2
    bold_row(r, 'Total Debits (₹)',  f"='All Transactions'!E{total_row}", DEB_FILL); r += 1
    bold_row(r, 'Total Credits (₹)', f"='All Transactions'!F{total_row}", CRD_FILL); r += 1
    bold_row(r, 'Net Spend (₹)',     f"='All Transactions'!E{total_row}-'All Transactions'!F{total_row}", SUM_FILL); r += 2
    hdr_cell(ws2, r, 1, 'Source File'); hdr_cell(ws2, r, 2, 'Transactions'); r += 1
    for src, cnt in Counter(t['source'] for t in sorted_txns).items():
        sum_row(r, src, cnt, fmt='0'); r += 1

    # Source Details sheet
    ws3 = wb.create_sheet('Source Details')
    for col, w in zip('ABCD', [35, 22, 22, 22]):
        ws3.column_dimensions[col].width = w
    ws3.merge_cells('A1:D1')
    t3 = ws3['A1']
    t3.value = 'Source File Details'
    t3.font = Font(name='Arial', bold=True, size=13, color='1F3864')
    t3.alignment = Alignment(horizontal='center')
    ws3.row_dimensions[1].height = 28
    for col, label in enumerate(['File Name', 'Card Number', 'Period From', 'Period To'], 1):
        hdr_cell(ws3, 2, col, label)
    for i, info in enumerate(account_infos, 3):
        for col, key in enumerate(['source','card_number','period_from','period_to'], 1):
            c = ws3.cell(row=i, column=col, value=info.get(key, '-'))
            c.font = BODY_FONT; c.border = thin_border
            c.fill = ALT_FILL if i % 2 == 0 else PatternFill()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/unlock', methods=['POST'])
def unlock():
    password = request.form.get('password', '')
    session_id = request.form.get('session_id', str(uuid.uuid4()))
    files = request.files.getlist('pdfs')

    session = SESSIONS.setdefault(session_id, {'unlocked': {}, 'locked': {}})

    # Merge previously locked + newly uploaded into pending
    pending = dict(session.get('locked', {}))  # fname -> tmp_path
    for f in files:
        if f.filename:
            path = save_to_tmp(f.read())
            pending[f.filename] = path

    if not pending and not session['unlocked']:
        return jsonify({'error': 'No files uploaded'}), 400

    new_locked = {}
    unlocked_names = []
    locked_names = []

    for filename, tmp_path in pending.items():
        try:
            with open(tmp_path, 'rb') as fh:
                data = fh.read()
            reader = PdfReader(io.BytesIO(data))
            if reader.is_encrypted:
                result = reader.decrypt(password)
                if result == 0:
                    new_locked[filename] = tmp_path
                    locked_names.append(filename)
                    del data
                    continue
            writer = PdfWriter()
            for page in reader.pages:
                writer.add_page(page)
            buf = io.BytesIO()
            writer.write(buf)
            new_path = save_to_tmp(buf.getvalue())
            delete_tmp(tmp_path)
            session['unlocked'][filename] = new_path
            unlocked_names.append(filename)
            del data, buf
        except Exception:
            new_locked[filename] = tmp_path
            locked_names.append(filename)

    # Clean up old locked tmp files that are no longer pending
    for fname, old_path in session.get('locked', {}).items():
        if fname not in new_locked:
            delete_tmp(old_path)

    session['locked'] = new_locked
    gc.collect()

    return jsonify({
        'session_id': session_id,
        'unlocked': unlocked_names,
        'locked': locked_names,
        'total_unlocked': len(session['unlocked']),
    })

@app.route('/consolidate', methods=['POST'])
def consolidate():
    session_id = request.form.get('session_id', '')
    files = request.files.getlist('pdfs')

    # Build map of fname -> (path, we_own_it)
    pdf_path_map = {}

    if session_id and session_id in SESSIONS:
        for fname, path in SESSIONS[session_id].get('unlocked', {}).items():
            pdf_path_map[fname] = (path, False)

    for f in files:
        if f.filename:
            raw = f.read()
            try:
                reader = PdfReader(io.BytesIO(raw))
                if reader.is_encrypted:
                    return jsonify({'error': f'{f.filename} is still password-protected. Please unlock first.'}), 400
            except Exception:
                pass
            path = save_to_tmp(raw)
            del raw
            pdf_path_map[f.filename] = (path, True)

    if not pdf_path_map:
        return jsonify({'error': 'No PDFs to consolidate'}), 400

    all_transactions = []
    account_infos = []
    parse_errors = []

    # Process files ONE AT A TIME — never hold all PDFs in RAM simultaneously
    for filename, (pdf_path, owned) in pdf_path_map.items():
        try:
            info, txns = parse_pdf_from_path(pdf_path, filename)
            account_infos.append(info)
            all_transactions.extend(txns)
        except Exception as e:
            parse_errors.append(f'{filename}: {str(e)}')
        finally:
            if owned:
                delete_tmp(pdf_path)
            gc.collect()

    if not all_transactions:
        return jsonify({
            'error': 'Could not extract any transactions. PDFs may be scanned images or use an unsupported format.',
            'parse_errors': parse_errors,
        }), 422

    total_count = len(all_transactions)
    files_count = len(pdf_path_map)

    xlsx_bytes = build_excel(all_transactions, account_infos)
    del all_transactions, account_infos
    gc.collect()

    # Save Excel to disk, not RAM
    dl_id = str(uuid.uuid4())
    fd, xlsx_path = tempfile.mkstemp(suffix='.xlsx')
    with os.fdopen(fd, 'wb') as f:
        f.write(xlsx_bytes)
    del xlsx_bytes
    SESSIONS[dl_id] = {'xlsx_path': xlsx_path}

    return jsonify({
        'download_id': dl_id,
        'total_transactions': total_count,
        'files_processed': files_count,
        'parse_errors': parse_errors,
    })

@app.route('/download/excel/<dl_id>')
def download_excel(dl_id):
    if dl_id not in SESSIONS or 'xlsx_path' not in SESSIONS[dl_id]:
        return jsonify({'error': 'Not found'}), 404
    xlsx_path = SESSIONS[dl_id]['xlsx_path']

    def generate():
        with open(xlsx_path, 'rb') as f:
            while True:
                chunk = f.read(64 * 1024)
                if not chunk:
                    break
                yield chunk
        delete_tmp(xlsx_path)
        SESSIONS.pop(dl_id, None)

    return Response(
        generate(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=SBI_Consolidated_Statement.xlsx'}
    )

@app.route('/download/unlocked/<session_id>')
def download_unlocked(session_id):
    if session_id not in SESSIONS:
        return jsonify({'error': 'Session not found'}), 404
    unlocked = SESSIONS[session_id].get('unlocked', {})
    if not unlocked:
        return jsonify({'error': 'No unlocked files'}), 404

    if len(unlocked) == 1:
        fname, path = next(iter(unlocked.items()))
        return send_file(path, mimetype='application/pdf',
                         as_attachment=True, download_name=f'unlocked_{fname}')

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fname, path in unlocked.items():
            zf.write(path, f'unlocked_{fname}')
    zip_buf.seek(0)
    return send_file(zip_buf, mimetype='application/zip',
                     as_attachment=True, download_name='unlocked_pdfs.zip')

@app.route('/clear/<session_id>', methods=['POST'])
def clear(session_id):
    cleanup_session(session_id)
    return jsonify({'ok': True})

if __name__ == '__main__':
    app.run(debug=True, port=5000)
